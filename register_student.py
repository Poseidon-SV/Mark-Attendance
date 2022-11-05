import cv2 as cv, os
from openpyxl import load_workbook

# pyinstaller --onefile pythonScriptName.py

haar_file = 'haarcascade_frontalface_default.xml'
database = 'database'
personDB = 'students'

confirm = False                                         # Confirming person's name
while confirm == False:
    confirm_name = input("Please enter your full name: ")
    face_of = input("Confirm your name: ")

    if face_of == confirm_name:
        confirm = True
    else:
        print("Name doesn't match. Please try again... \n")


path = os.path.join(database,face_of)                   # Making database of that person
if not os.path.isdir(path):
    os.makedirs(path)
pathS = os.path.join(personDB,face_of)                   # Making database of that person
if not os.path.isdir(pathS):
    os.makedirs(pathS)

width, height = 130, 100

face_cascade = cv.CascadeClassifier(haar_file)          # Making sure it's face photo
cam = cv.VideoCapture(0)

count = 0   
while count < 61: # To keep track of photos (max 30)
    _, img = cam.read()
    if count == 20 or count == 40:
        cv.imwrite('%s/%s.png' % (pathS, count), img)
    print("Photo no.",count) 
    status = f"Photo collected: {count}"
    gray = cv.cvtColor(img, cv.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.3, 4)
    for (x, y, w, h) in faces:
        cv.rectangle(img,(x, y), (x+w, y+h), (255, 0, 0), 2)
        face = gray[y:y + h, x:x + w]
        face_resize = cv.resize(face, (width, height))
        cv.imwrite('%s/%s.png' % (path, count), face_resize)
        status = f"Photo collected: {count}"
        count += 1
        cv.putText(img, status, (10, 20), cv.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255) if count < 30 else (0, 255, 0), 2)
    
    cv.putText(img, status, (10, 20), cv.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255) if count < 30 else (0, 255, 0), 2)        
    cv.imshow('Capturing Data', img)
    

    key = cv.waitKey(10)
    if key == 27:
        break        


workbook = load_workbook(filename="Attendance.xlsx")
sheet = workbook.active
r = 1
c = 1
present = sheet.cell(row = r, column = c)
student = present.value
while type(student) == str:
    present = sheet.cell(row = r, column = c)
    student = present.value    
    r = r + 1
    if student == None:
        sheet.cell(row = r-1, column = c).value = confirm_name
workbook.save(filename="Attendance.xlsx")


# applicationForm = load_workbook(filename="Application_form.xlsx")
# sheetForm = applicationForm.active
# studentList = []
# for i in range(2, (sheetForm.max_row+1)):
#     # if sheetForm.cell(row = i, column = 3).value != confirm_name:
#     studentA = sheetForm.cell(row = i, column = 3).value
#     studentA = studentA.replace(' ','')
#     studentList.append(studentA)
# c = confirm_name.replace(' ','')
# if c not in studentList:
#     import winsound
#     frequency = 2000
#     duration = 2000
#     winsound.Beep(frequency, duration)
#     print("Fill the Application Form", confirm_name,'!!!')


print("Data collected and Attendace marked :)")
cam.release()
cv.destroyAllWindows()  

