import cv2 as cv, numpy, os
from openpyxl import load_workbook
from itertools import groupby

haar_files = "haarcascade_frontalface_default.xml"
database = 'database'

(images, labels, names, id) = ([], [], {}, 0)
for (subdirs, dirs, files) in os.walk(database):
    for subdir in dirs:
        names[id] = subdir 
        subjectpath = os.path.join(database, subdir)
        for filename in os.listdir(subjectpath):
            path = subjectpath + '/' + filename
            label = id
            images.append(cv.imread(path, 0))
            labels.append(int(label))
        id += 1

(width, height) = (130, 100)
(images,labels) = [numpy.array(lis) for lis in [images, labels]]

model = cv.face.LBPHFaceRecognizer_create()
# OR
# model = cv.face.FisherFaceRecognizer_create()
model.train(images,labels)

face_cascade = cv.CascadeClassifier(haar_files)
webcam = cv.VideoCapture(0)
cnt = 0

c = 2 # Day
workbook = load_workbook(filename="Attendance.xlsx")
sheet = workbook.active
count = 0
confirm = []
confimAttendance =[]
while True:
    (_, img) = webcam.read()
    gray = cv.cvtColor(img, cv.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.3, 5)
    for (x,y,w,h) in faces:
        cv.rectangle(img, (x,y), (x+w, y+h), (255,255,0),2)
        face = gray[y:y+h, x:x + w]
        face_resize = cv.resize(face, (width, height))

        prediction = model.predict(face_resize) # Face recognition starts from here
        cv.rectangle(img, (x, y), (x + w, y + h), (0,255,0),3)
        if prediction[1]<80:   # Format float with no decimal places        # ( name of person, Confidence of person ) 
            cv.putText(img, '%s - %.0f' % (names[prediction[0]], prediction[1]), (x-10, y-10), cv.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2) # The %s operator is put where the string is to be specified. print("Hey, %s!" % name)
            # cv.putText(img, 'Attendance marked', (10, 20), cv.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255),2) 
            # print(names[prediction[0]],'Attendance marked')
            # cnt = 0
            # count += 1
            confirm.append(names[prediction[0]])     
            # print(confirm)       

            # print(count)
            if any(sum(1 for _ in g) > 10 for _, g in groupby(confirm)):
                print(names[prediction[0]],'Attendance marked')               
                cv.putText(img, 'Attendance marked', (10, 20), cv.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0),3) 
                for i in range(1, (sheet.max_row+1)):
                    if sheet.cell(row = i, column = 1).value == names[prediction[0]]:
                        sheet.cell(row = i, column = c).value = "P"
                confimAttendance = confirm
                confirm = []
            else: 
                if names[prediction[0]] in confimAttendance :
                    cv.putText(img, 'Attendance marked', (10, 20), cv.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0),3) 
                else:
                    cv.putText(img, 'confirming....', (10, 20), cv.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255),3) 
                    print('confirming....')

            # sheet.cell(row = r, column = c).value = "P"
        else:
            cnt+=1
            cv.putText(img, 'Unknown', (x-10, y-10), cv.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)       
            if (cnt>100):
                print('Unknown Person')
                cv.imwrite("input.jpg", img) # Takes and save the photo of unkown person if detected
                cnt = 0

    cv.imshow('Face Recognition', img)
    key = cv.waitKey(10)
    if key == 27:
        break

workbook.save(filename="Attendance.xlsx")
webcam.release()
cv.destroyAllWindows()
